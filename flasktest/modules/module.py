from flasktest.database import database, connect_db, disconnect_db
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Border, Font, Side
import openpyxl
import json


class Contingent:
    def __init__(self, edu, sm, faculty_id, profession_id, sdate, edate, radio) -> None:
        self.edu = edu
        self.sm = sm
        self.faculty_id = faculty_id
        self.profession_id = profession_id
        self.sdate = sdate
        self.edate = edate
        self.radio = radio
        self.table_end = 0
        self.db_name = self._get_db_name(self.faculty_id)
        self.coordinates = self.read_data()
        self.get_with_profession = "-- "
        if self.profession_id:
            self.get_with_profession = ""

        self.workbook = openpyxl.Workbook()
        self.ws = self.workbook.active
        # Naxçıvan Dövlət Universitetinin  _əyani___ şöbəsinin
        #   _Pedaqoji_fakültəsi  üzrə təhsil alan tələbə kontingentinin dəyişməsi haqqında
        self.merge_cells("A1:AG1")
        self.ws[f"A1"].value = "Naxçıvan Dövlət Universitetinin əyani şöbəsinin"
        self.ws[f"A1"].alignment = Alignment(horizontal="center", vertical="center")
        self.ws[f"A1"].font = Font(bold=True)

        self.merge_cells("A2:AG2")
        faculty_name = self._get_faculty_name(self.faculty_id)
        self.ws[f"A2"].value = (
            f"{faculty_name} fakültəsi üzrə təhsil alan tələbə kontingentinin dəyişməsi haqqında"
        )
        self.ws[f"A2"].alignment = Alignment(horizontal="center", vertical="center")
        self.ws[f"A2"].font = Font(bold=True)

        self.merge_cells("AD4:AG4")
        self.ws[f"AD4"].value = f"{self.sdate} - {self.edate}"
        self.ws[f"AD4"].alignment = Alignment(horizontal="right", vertical="center")
        self.ws[f"AD4"].font = Font(bold=True)

        self.merge_cells("A6:AG6")
        if self.profession_id:
            text = self._get_profession_name(self.profession_id)
        else:
            text = "fakültə üzrə"
        self.ws[f"A6"].value = text
        self.ws[f"A6"].alignment = Alignment(horizontal="right", vertical="center")
        self.ws[f"A6"].font = Font(bold=True)

        self.create_headings()
        result_dict = self.get_students()
        result_dict2, ak_mez = self.get_students_after_movements()

        self.create_student_count(result_dict, result_dict2, ak_mez)

        data = self.get_contingent_movements()
        sorted_data = {}
        for i in data:
            category_name = (
                i["income_category"]
                if i["gone_category"] is None
                else i["gone_category"]
            )
            if i["income_category"] in sorted_data or i["gone_category"] in sorted_data:

                if i["foreign"]:
                    sorted_data[category_name][i["course"]]["xarici"] += 1
                if i["o/d"]:
                    sorted_data[category_name][i["course"]]["odenisli"] += 1
                else:
                    sorted_data[category_name][i["course"]]["odenissiz"] += 1
            else:
                sorted_data[category_name] = {}

                sorted_data[category_name][i["course"]] = {
                    "odenisli": 0,
                    "odenissiz": 0,
                    "xarici": 0,
                }
                if i["foreign"]:
                    sorted_data[category_name][i["course"]]["xarici"] += 1
                if i["o/d"]:
                    sorted_data[category_name][i["course"]]["odenisli"] += 1
                else:
                    sorted_data[category_name][i["course"]]["odenissiz"] += 1

        self.create_table(sorted_data)

    def merge_cells(self, range: str):
        self.ws.merge_cells(range)

    def __format_ws__(self, cell_range):
        # applying border and alignment
        font = Font(size=10)
        # align = Alignment(vertical='center', horizontal='left', indent=1, wrapText=True)
        # border = Border(left=Side(border_style='thin', color='000000'),
        #                 right=Side(
        #                     border_style='thin', color='000000'),
        #                 top=Side(
        #                     border_style='thin', color='000000'),
        #                 bottom=Side(border_style='thin', color='000000'))
        border = Border(
            left=Side(border_style="thin", color="ffa0a0a0"),
            right=Side(border_style="thin", color="ffa0a0a0"),
            top=Side(border_style="thin", color="ffa0a0a0"),
            bottom=Side(border_style="thin", color="ffa0a0a0"),
        )

        rows = [rows for rows in self.ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [
            (setattr(cell, "border", border), setattr(cell, "font", font))
            for cell in flattened
        ]

    def create_headings(self):
        for column in self.coordinates["columns"]:
            self.merge_cells(column["range"])
            self.ws[column["key"]].value = column["text"]
            self.ws[column["key"]].alignment = Alignment(
                horizontal="center", vertical="center", **column["kwargs"]
            )
        # ------------------------------------------------------------------------------------
        for column in self.coordinates["sub_columns"]:
            self.merge_cells(column["range"])
            self.ws[column["key"]].value = column["text"]
            self.ws[column["key"]].alignment = Alignment(
                horizontal="center", vertical="center", **column["kwargs"]
            )

    def create_student_count(self, result_dict, result_dict2, ak_mez):
        totals = {"odenisli": 0, "odenissiz": 0, "xarici": 0}
        totals2 = {
            "odenisli": 0,
            "xarici": 0,
            "odenisli_kisi": 0,
            "odenisli_qadin": 0,
            "odenissiz": 0,
            "odenissiz_kisi": 0,
            "odenissiz_qadin": 0,
            "xarici_odenisli_kisi": 0,
            "xarici_odenisli_qadin": 0,
        }
        totals3 = {"odenisli_akmez": 0, "xarici_akmez": 0, "odenissiz_akmez": 0}

        start = 17
        for i in result_dict:
            print(result_dict[i])
            print(result_dict2[i])
            print(ak_mez[i])
            self.create_course(i, result_dict[i], result_dict2[i], ak_mez[i], start, 0)
            start += 4

            for a in result_dict[i]:
                totals[a] += result_dict[i][a]

            totals2["odenisli"] += result_dict2[i]["odenisli"]
            totals2["xarici"] += result_dict2[i]["xarici"]
            totals2["odenisli_kisi"] += result_dict2[i]["odenisli_kisi"]
            totals2["odenisli_qadin"] += result_dict2[i]["odenisli_qadin"]
            totals2["odenissiz"] += result_dict2[i]["odenissiz"]
            totals2["odenissiz_kisi"] += result_dict2[i]["odenissiz_kisi"]
            totals2["odenissiz_qadin"] += result_dict2[i]["odenissiz_qadin"]
            totals2["xarici_odenisli_kisi"] += result_dict2[i]["xarici_odenisli_kisi"]
            totals2["xarici_odenisli_qadin"] += result_dict2[i]["xarici_odenisli_qadin"]

            totals3["odenisli_akmez"] += ak_mez[i]["odenisli_akmez"]
            totals3["xarici_akmez"] += ak_mez[i]["xarici_akmez"]
            totals3["odenissiz_akmez"] += ak_mez[i].get("odenissiz_akmez", 0)

        self.create_course("Ümumi cəmi", totals, totals2, totals3, start, 90)
        self.table_end = start + 3

    def create_course(self, key, values, values2, ak_mez, start, rotation):
        if not values2.get("odenissiz", False):
            values2["odenissiz"] = 0
            values2["odenissiz_qadin"] = 0
            values2["odenissiz_kisi"] = 0

        converter = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI"}
        self.merge_cells(f"A{start}:A{start+3}")
        self.ws[f"A{start}"].value = converter.get(key, "Ümumi cəmi")
        self.ws[f"A{start}"].alignment = Alignment(
            horizontal="center", vertical="center", textRotation=rotation
        )

        for index, i in enumerate(["Ödənişli", "Ödənişsiz", "Xarici", "Cəmi"]):
            self.ws[f"B{start+index}"].value = i
            self.ws[f"B{start+index}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        myKeys = list(values.keys())
        myKeys.sort()
        sorted_dict = {i: values[i] for i in myKeys}

        sum_c = sum(sorted_dict.values())
        sum_ak_mez = sum(ak_mez.values())
        for index, (key, value) in enumerate(sorted_dict.items()):
            if value != 0:
                self.ws[f"C{start+index}"].value = value
                self.ws[f"C{start+index}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            if index == 0:
                key1 = "odenisli"
                key2 = "odenisli_kisi"
                key3 = "odenisli_qadin"
            elif index == 1:
                key1 = "odenissiz"
                key2 = "odenissiz_kisi"
                key3 = "odenissiz_qadin"
            else:
                key1 = "xarici"
                key2 = "xarici_odenisli_kisi"
                key3 = "xarici_odenisli_qadin"
            if values2[key1] != 0:
                self.ws[f"AD{start+index}"].value = values2[key1]
                self.ws[f"AD{start+index}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                self.ws[f"AE{start+index}"].value = values2[key2]
                self.ws[f"AE{start+index}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                self.ws[f"AF{start+index}"].value = values2[key3]
                self.ws[f"AF{start+index}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if index == 0:
                key = "odenisli_akmez"
            elif index == 1:
                key = "xarici_akmez"
            else:
                key = "odenissiz_akmez"
            if ak_mez[key] != 0:
                self.ws[f"AG{start+index}"].value = ak_mez[key]
                self.ws[f"AG{start+index}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

        self.ws[f"C{start+3}"].value = sum_c
        self.ws[f"C{start+3}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        self.ws[f"AD{start+3}"].value = (
            values2["odenisli"] + values2["odenissiz"] + values2["xarici"]
        )
        self.ws[f"AD{start+3}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.ws[f"AE{start+3}"].value = (
            values2["odenisli_kisi"]
            + values2["odenissiz_kisi"]
            + values2["xarici_odenisli_kisi"]
        )
        self.ws[f"AE{start+3}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.ws[f"AF{start+3}"].value = (
            values2["odenisli_qadin"]
            + values2["odenissiz_qadin"]
            + values2["xarici_odenisli_qadin"]
        )
        self.ws[f"AF{start+3}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        if sum_ak_mez != 0:
            self.ws[f"AG{start+3}"].value = sum_ak_mez
            self.ws[f"AG{start+3}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

    def create_table(self, data):
        for category in data:
            for i in self.coordinates["sub_columns"]:
                if i["text"] == category:
                    letter = i["letter"]
                    break
            else:
                letter = None

            # course_to_row = {1: 17, 2: 21, 3: 25, 4: 29, 5: 33, 6: 37}
            converter = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI"}

            total = {"odenisli": 0, "odenissiz": 0, "xarici": 0}

            for course in data[category]:
                for index in [17, 21, 25, 29, 33, 37]:
                    if self.ws[f"A{index}"].value == converter[course]:
                        row = index
                        break
                else:
                    row = None

                print(self.ws[f"A{row}"].value, converter[course])

                column = data[category][course]
                odenisli = column["odenisli"]
                odenissiz = column["odenissiz"]
                xarici = column["xarici"]

                total["odenisli"] += odenisli
                total["odenissiz"] += odenissiz
                total["xarici"] += xarici

                cemi = odenisli + odenissiz + xarici

                if odenisli != 0:
                    self.ws[f"{letter}{row}"].value = odenisli
                    self.ws[f"{letter}{row}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                if odenissiz != 0:
                    self.ws[f"{letter}{row+1}"].value = odenissiz
                    self.ws[f"{letter}{row+1}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                if xarici != 0:
                    self.ws[f"{letter}{row+2}"].value = xarici
                    self.ws[f"{letter}{row+2}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                if cemi != 0:
                    self.ws[f"{letter}{row+3}"].value = cemi
                    self.ws[f"{letter}{row+3}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                # row += 4
            odenisli = total["odenisli"]
            odenissiz = total["odenissiz"]
            xarici = total["xarici"]
            cemi = odenisli + odenissiz + xarici

            row = self.table_end - 3
            if odenisli != 0:
                self.ws[f"{letter}{row}"].value = odenisli
                self.ws[f"{letter}{row}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if odenissiz != 0:
                self.ws[f"{letter}{row+1}"].value = odenissiz
                self.ws[f"{letter}{row+1}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if xarici != 0:
                self.ws[f"{letter}{row+2}"].value = xarici
                self.ws[f"{letter}{row+2}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if cemi != 0:
                self.ws[f"{letter}{row+3}"].value = cemi
                self.ws[f"{letter}{row+3}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    def read_data(self):
        with open("flasktest/modules/coordinates.json", encoding="utf-8") as f:
            data = json.loads(f.read())
        return data

    def save(self, name):
        self.__format_ws__(f"A6:AG{self.table_end}")
        self.workbook.save(f"excel-files/{name}.xlsx")

    def _get_db_name(self, faculty_id):
        connection = connect_db(database)
        with connection.cursor() as cursor:
            query = (
                f"SELECT db_name FROM examsystem.faculty_names where id={faculty_id};"
            )
            cursor.execute(query)
            result = cursor.fetchall()

        disconnect_db(connection, database)

        if result:
            return result[0][0]
        return None

    def _get_faculty_name(self, faculty_id):
        connection = connect_db(database)
        with connection.cursor() as cursor:
            query = f"SELECT faculty_name FROM examsystem.faculty_names where id={faculty_id};"
            cursor.execute(query)
            result = cursor.fetchall()

        disconnect_db(connection, database)

        if result:
            return result[0][0]
        return None

    def _get_profession_name(self, profession_id):
        connection = connect_db(database)
        with connection.cursor() as cursor:
            query = f"SELECT profession_name FROM examsystem.professions where id={profession_id};"
            cursor.execute(query)
            result = cursor.fetchall()

        disconnect_db(connection, database)

        if result:
            return result[0][0]
        return None

    def get_students(self):
        connection = connect_db(database)
        with connection.cursor() as cursor:
            query = f"""SELECT 
                            std.course,
                            SUM(CASE
                                    WHEN foreign_student = FALSE THEN 1
                                    ELSE 0
                                END) AS normal,
                                SUM(CASE
                                    WHEN foreign_student = TRUE THEN 1
                                    ELSE 0
                                END) AS xarici
                        FROM
                            {self.db_name}.students as std
                                LEFT JOIN
                            examsystem.contingent_movements AS cm ON cm.student_id = std.id
                                JOIN
                            examsystem.professions AS pr on pr.id=std.profession_id
                        WHERE
                            educationYear = '{self.edu}'
                                AND semestr = '{self.sm}'
                                {self.get_with_profession} AND std.profession_id={self.profession_id}
                                AND foreign_student = FALSE
                                AND pr.sectors={self.radio}
                                AND (cm.date>="{self.edate}" OR cm.date is Null)
                                OR (cm.date <= '{self.sdate}'
                                AND cm.incomers_action IS NOT NULL)
                                
                                
                        GROUP BY course , `o/d`
                        ORDER BY course, `o/d` desc;
                    """
            print(query)
            cursor.execute(query)
            result = cursor.fetchall()

            result_dict = {}
            for key, normal, foreign in result:
                if key in result_dict:
                    result_dict[key]["odenissiz"] = int(normal)
                else:
                    result_dict[key] = {}
                    result_dict[key]["odenisli"] = int(normal)
                    result_dict[key]["xarici"] = int(foreign)

        disconnect_db(connection, database)
        return result_dict

    def get_contingent_movements(self):

        connection = connect_db(database)
        data = []
        with connection.cursor() as cursor:
            query = f"""
                    SELECT 
                        cm.student_id, ci.category_name, cg.category_name
                    FROM
                        examsystem.contingent_movements AS cm
                            LEFT JOIN
                        examsystem.contingent_incomers AS ci ON ci.id = cm.incomers_action
                            LEFT JOIN
                        examsystem.contingent_goners AS cg ON cg.id = cm.goners_action
                            JOIN
                        examsystem.professions AS pr on pr.id=cm.profession_id

                    WHERE
                        cm.faculty_id = {self.faculty_id} AND (date >= '{self.sdate}' and date <= "{self.edate}")
                        AND pr.sectors={self.radio}
                        {self.get_with_profession} AND cm.profession_id={self.profession_id}
                    """
            cursor.execute(query)
            result = cursor.fetchall()

            for i in result:
                query = f"""
                        SELECT * from {self.db_name}.students where id={i[0]};
                        """
                cursor.execute(query)
                student_result = cursor.fetchall()
                temp = {
                    "income_category": i[1],
                    "gone_category": i[2],
                    "course": student_result[0][4],
                    "foreign": student_result[0][9],
                    "o/d": student_result[0][11],
                }
                data.append(temp)

        disconnect_db(connection, database)
        return data

    def get_students_after_movements(self):
        connection = connect_db(database)
        with connection.cursor() as cursor:
            query = f"""
                        SELECT 
                            std.course,
                            SUM(CASE
                                WHEN foreign_student = FALSE THEN 1
                                ELSE 0
                            END) AS normal,
                            SUM(CASE
                                WHEN foreign_student = TRUE THEN 1
                                ELSE 0
                            END) AS xarici,
                            SUM(CASE
                                WHEN
                                    gender = TRUE
                                        AND foreign_student = FALSE
                                THEN
                                    1
                                ELSE 0
                            END) AS yerli_kisi,
                            SUM(CASE
                                WHEN
                                    gender = FALSE
                                        AND foreign_student = FALSE
                                THEN
                                    1
                                ELSE 0
                            END) AS yerli_qadin,
                            SUM(CASE
                                WHEN
                                    gender = TRUE
                                        AND foreign_student = True
                                THEN
                                    1
                                ELSE 0
                            END) AS xarici_kisi,
                            SUM(CASE
                                WHEN
                                    gender = FALSE
                                        AND foreign_student = True
                                THEN
                                    1
                                ELSE 0
                            END) AS xarici_qadin
                        FROM
                            {self.db_name}.students AS std
                                LEFT JOIN
                            examsystem.contingent_movements AS cm ON cm.student_id = std.id
                                JOIN
                            examsystem.professions AS pr on pr.id=std.profession_id
                        WHERE
                            educationYear = '{self.edu}'
                                AND semestr = '{self.sm}'
                                AND pr.sectors = {self.radio}
                                {self.get_with_profession} AND std.profession_id={self.profession_id}
                                AND is_active = 1
                                
                                
                        GROUP BY course , `o/d`
                        ORDER BY course , `o/d` DESC;
                    """
            print(query)
            cursor.execute(query)
            result = cursor.fetchall()

            result_dict = {}
            for key, normal, foreign, man, woman, foreign_man, foreign_woman in result:
                if key in result_dict:
                    result_dict[key]["odenissiz"] = int(normal)
                    result_dict[key]["odenissiz_kisi"] = int(man)
                    result_dict[key]["odenissiz_qadin"] = int(woman)
                else:
                    result_dict[key] = {}
                    result_dict[key]["odenisli"] = int(normal)
                    result_dict[key]["xarici"] = int(foreign)
                    result_dict[key]["odenisli_kisi"] = int(man)
                    result_dict[key]["odenisli_qadin"] = int(woman)
                    result_dict[key]["xarici_odenisli_kisi"] = int(foreign_man)
                    result_dict[key]["xarici_odenisli_qadin"] = int(foreign_woman)

            query = f"""
                        SELECT 
                            std.course,
                            SUM(CASE
                                WHEN cm.goners_action IN (1 , 2, 4) THEN 1
                                ELSE 0
                            END) AS `ak.mez`,
                            SUM(CASE
                                WHEN
                                    foreign_student = TRUE
                                        AND cm.goners_action IN (1 , 2, 4)
                                THEN
                                    1
                                ELSE 0
                            END) AS xarici
                        FROM
                            {self.db_name}.students AS std
                                LEFT JOIN
                            examsystem.contingent_movements AS cm ON cm.student_id = std.id
                                JOIN
                            examsystem.professions AS pr on pr.id=std.profession_id
                        WHERE 
                            pr.sectors = {self.radio}
                            {self.get_with_profession} AND std.profession_id={self.profession_id}
                        GROUP BY course , `o/d`
                        ORDER BY course , `o/d` DESC;
                    """
            cursor.execute(query)
            ak_mez = cursor.fetchall()
            result_ak_mez = {}
            for key, akmez, foreign in ak_mez:
                if key in result_ak_mez:
                    result_ak_mez[key]["odenissiz_akmez"] = int(akmez)
                    result_ak_mez[key]["xarici_akmez"] += int(foreign)
                else:
                    result_ak_mez[key] = {}
                    result_ak_mez[key]["odenisli_akmez"] = int(akmez)
                    result_ak_mez[key]["xarici_akmez"] = int(foreign)

        disconnect_db(connection, database)
        return result_dict, result_ak_mez


class MovementReport:
    def __init__(self, data, faculty_id) -> None:
        self.faculty_id = faculty_id
        self.data = self.filter_data(data)
        self.workbook = openpyxl.Workbook()
        self.ws = self.workbook.active
        self.put_table()

    def filter_data(self, data):
        if self.faculty_id == 0:
            return [i[:-3] for i in data]

        data = [i[:-3] for i in data if i[10] == self.faculty_id]
        return data

    def put_table(self):
        columns = [
            "Fakültə adı",
            "İxtisas adı",
            "Kurs",
            "Tələbə Ad",
            "Gəlməşdir",
            "Getmişdir",
            "Tarix",
            "Əmr nömrəsi",
        ]
        self.ws.append(columns)
        for row in self.data:
            self.ws.append(row)

    def save(self, name):
        self.workbook.save(f"excel-files/{name}.xlsx")
